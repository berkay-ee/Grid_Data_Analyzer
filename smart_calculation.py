import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def browse_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")])
    entry_path.delete(0, tk.END)
    entry_path.insert(0, file_path)

def split_excel():
    file_path = entry_path.get()
    if not file_path:
        messagebox.showerror("Error", "Please select an Excel file!")
        return

    try:
        df = pd.read_excel(file_path)
        abone_numbers = df['Abone No'].unique()

        # Klasör oluştur
        folder_name = "Abone_Files"
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        for abone in abone_numbers:
            abone_df = df[df['Abone No'] == abone].copy()

            numeric_cols = abone_df.select_dtypes(include=['int64', 'float64']).columns

            # Summary row
            summary_row = abone_df[numeric_cols].sum()
            summary_row["Abone No"] = abone
            summary_row["Ünvan"] = "TOPLAM"
            summary_row["Tarih"] = ""

            abone_df = pd.concat([abone_df, summary_row.to_frame().T], ignore_index=True)

            # Export Excel
            output_file = os.path.join(folder_name, f"{abone}.xlsx")
            abone_df.to_excel(output_file, index=False)

            # TOPLAM satırını renklendirme
            wb = load_workbook(output_file)
            ws = wb.active
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill
            wb.save(output_file)

        messagebox.showinfo("Success", f"Split complete! {len(abone_numbers)} files created in '{folder_name}' folder.")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# GUI setup
root = tk.Tk()
root.title("Excel Splitter by Abone No")
root.geometry("500x150")

tk.Label(root, text="Excel file path:").pack(pady=5)
entry_path = tk.Entry(root, width=60)
entry_path.pack(pady=5)

tk.Button(root, text="Browse", command=browse_file).pack(pady=5)
tk.Button(root, text="Split Excel by Abone No", command=split_excel).pack(pady=10)

root.mainloop()
